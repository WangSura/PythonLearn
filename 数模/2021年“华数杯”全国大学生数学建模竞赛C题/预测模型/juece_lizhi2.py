# 1.读取数据与简单预处理
import openpyxl
import pandas as pd
import joblib
import os

dirs = 'D:/program/pythonPractice/数模/2021年“华数杯”全国大学生数学建模竞赛C题/预测模型'
if not os.path.exists(dirs):
    os.makedirs(dirs)

# 读取模型
model = joblib.load(dirs+'/model2.pkl')
# wb = openpyxl.load_workbook(
#    'D:/program/pythonPractice/数模/2021年“华数杯”全国大学生数学建模竞赛C题/预测模型/待判定的数据2.xlsx')
#wb_features = wb[wb.columns[2:9]]
#wb_list = list(wb_features.columns)
# print(wb.max_row)
# print(wb.max_column)
# for i in range(3, 10, 1):
#sh_name = wb.get_sheet_names()
#ws = wb[wb.sheetnames[0]]
#print(ws.cell(2, i).value)
#ws.cell(2, i).value = (1+5/100)*ws.cell(2, i).value
#print(ws.cell(2, i).value)
# wb.save('D:/program/pythonPractice/数模/2021年“华数杯”全国大学生数学建模竞赛C题/预测模型/待判定的数据22.xlsx')
# i++
test = pd.read_excel(
    'D:/program/pythonPractice/数模/2021年“华数杯”全国大学生数学建模竞赛C题/预测模型/待判定的数据22.xlsx')
test = test[test.columns[0:27]]
print('预测结果:\n', model.predict(test))
